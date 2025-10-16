from __future__ import annotations

import json
from datetime import datetime
from typing import Any, Iterable
from io import BytesIO

from flask import Blueprint, jsonify, request, send_file

from ..db import get_connection
from ..schemas import BudgetIncreaseDecision, SolicitudCreate, SolicitudDraft
from ..security import verify_access_token

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


bp = Blueprint("solicitudes", __name__, url_prefix="/api")

COOKIE_NAME = "spm_token"
STATUS_PENDING = "pendiente_de_aprobacion"
STATUS_APPROVED = "aprobada"
STATUS_REJECTED = "rechazada"
STATUS_CANCELLED = "cancelada"
STATUS_FINALIZED = "finalizada"
STATUS_DRAFT = "draft"
STATUS_CANCEL_PENDING = "cancelacion_pendiente"
STATUS_CANCEL_REJECTED = "cancelacion_rechazada"
STATUS_IN_TREATMENT = "en_tratamiento"


def _utcnow_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def _get_auth_token() -> str | None:
    token = request.cookies.get(COOKIE_NAME)
    if not token:
        header = request.headers.get("Authorization", "")
        if header.startswith("Bearer "):
            token = header.split(" ", 1)[1].strip()
    return token or None


def _require_auth() -> str | None:
    token = _get_auth_token()
    if not token:
        return None
    try:
        payload = verify_access_token(token)
    except Exception:
        return None
    sub = payload.get("sub")
    return str(sub).strip() if sub else None


def _json_error(code: str, message: str, status: int = 400):
    return jsonify({"ok": False, "error": {"code": code, "message": message}}), status


def _coerce_str(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _fetch_user(con, uid: str | None):
    if not uid:
        return None
    return con.execute(
        """
        SELECT id_spm, nombre, apellido, rol, centros, jefe, gerente1, gerente2
          FROM usuarios
         WHERE lower(id_spm)=?
        """,
        (uid.lower(),),
    ).fetchone()


def _normalize_uid(value: Any) -> str | None:
    normalized = _coerce_str(value).lower()
    return normalized or None


def _ensure_user_exists(con, uid: str | None) -> str | None:
    """Return a normalized user id only if it exists in usuarios."""
    normalized = _normalize_uid(uid)
    if not normalized:
        return None
    row = con.execute(
        "SELECT 1 FROM usuarios WHERE lower(id_spm)=?",
        (normalized,),
    ).fetchone()
    return normalized if row else None


def _has_role(user: dict[str, Any] | None, *needles: str) -> bool:
    if not user:
        return False
    role = _coerce_str(user.get("rol")).lower()
    for needle in needles:
        if needle.lower() in role:
            return True
    return False


def _resolve_approver(con, user: dict[str, Any] | None, total_monto: float = 0.0) -> str | None:
    if not user:
        return None
    
    # Determinar el aprobador basado en el monto total
    if total_monto <= 20000.0:
        # Jefe desde USD 0.01 hasta USD 20000
        approver_field = "jefe"
    elif total_monto <= 100000.0:
        # Gerente1 desde USD 20000.01 hasta USD 100000
        approver_field = "gerente1"
    else:
        # Gerente2 desde USD 100000.01 en adelante
        approver_field = "gerente2"
    
    approver_email = _coerce_str(user.get(approver_field))
    if approver_email:
        # Buscar el id_spm del usuario con este email
        approver_user = con.execute(
            "SELECT id_spm FROM usuarios WHERE lower(mail) = ?",
            (approver_email.lower(),)
        ).fetchone()
        if approver_user:
            return approver_user["id_spm"]
    
    # Fallback: buscar en otros campos si el campo específico no está disponible
    for field in ("jefe", "gerente1", "gerente2"):
        approver_email = _coerce_str(user.get(field))
        if approver_email:
            approver_user = con.execute(
                "SELECT id_spm FROM usuarios WHERE lower(mail) = ?",
                (approver_email.lower(),)
            ).fetchone()
            if approver_user:
                return approver_user["id_spm"]
    return None


def _resolve_planner(user: dict[str, Any] | None) -> str | None:
    if not user:
        return None
    for field in ("gerente2", "gerente1"):
        value = _coerce_str(user.get(field))
        if value:
            return value.lower()
    return None


def _normalize_items(raw_items: Iterable[Any]) -> tuple[list[dict[str, Any]], float]:
    items: list[dict[str, Any]] = []
    total = 0.0
    for raw in raw_items or []:
        if not isinstance(raw, dict):
            continue
        codigo = _coerce_str(raw.get("codigo"))
        if not codigo:
            continue
        descripcion = _coerce_str(raw.get("descripcion"))
        try:
            cantidad = int(raw.get("cantidad", 0))
        except (TypeError, ValueError):
            cantidad = 0
        if cantidad < 1:
            cantidad = 1
        precio_raw = raw.get("precio_unitario")
        if precio_raw is None:
            precio_raw = raw.get("precio")
        try:
            precio = float(precio_raw)
        except (TypeError, ValueError):
            precio = 0.0
        if precio < 0:
            precio = 0.0
        subtotal = round(cantidad * precio, 2)
        item: dict[str, Any] = {
            "codigo": codigo,
            "descripcion": descripcion,
            "cantidad": cantidad,
            "precio_unitario": round(precio, 2),
            "comentario": raw.get("comentario"),
            "subtotal": subtotal,
        }
        unidad = raw.get("unidad") or raw.get("uom") or raw.get("unidad_medida")
        if unidad:
            item["unidad"] = _coerce_str(unidad)
        items.append(item)
        total += subtotal
    return items, round(total, 2)


def _parse_draft_payload(uid: str, payload: dict[str, Any]) -> dict[str, Any]:
    model = SolicitudDraft(**{**payload, "id_usuario": uid})
    data = model.model_dump()
    data["id_usuario"] = uid.lower()
    fecha = data.get("fecha_necesidad")
    if fecha:
        data["fecha_necesidad"] = fecha.isoformat()
    data["criticidad"] = data.get("criticidad") or "Normal"
    return data


def _parse_full_payload(uid: str, payload: dict[str, Any], *, expect_items: bool) -> dict[str, Any]:
    sanitized_items = []
    for item in payload.get("items", []) or []:
        if not isinstance(item, dict):
            continue
        sanitized_items.append(
            {
                "codigo": item.get("codigo"),
                "descripcion": item.get("descripcion"),
                "cantidad": item.get("cantidad"),
                "precio_unitario": item.get("precio_unitario"),
                "comentario": item.get("comentario"),
            }
        )
    payload_for_model = {**payload, "items": sanitized_items, "id_usuario": uid}
    model = SolicitudCreate(**payload_for_model)
    data = model.model_dump()
    data["id_usuario"] = uid.lower()
    fecha = data.get("fecha_necesidad")
    if fecha:
        data["fecha_necesidad"] = fecha.isoformat()
    data["criticidad"] = data.get("criticidad") or "Normal"
    items, total = _normalize_items(payload.get("items", []))
    if expect_items and not items:
        raise ValueError("Debe incluir al menos un ítem")
    data["items"] = items
    data["total_monto"] = total
    return data


def _json_load(raw: Any) -> dict[str, Any]:
    if isinstance(raw, dict):
        return dict(raw)
    if not raw:
        return {}
    try:
        return json.loads(raw)
    except Exception:
        return {}


def _serialize_items(items: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for raw in items or []:
        if not isinstance(raw, dict):
            continue
        cantidad = raw.get("cantidad")
        try:
            cantidad_int = int(cantidad)
        except (TypeError, ValueError):
            cantidad_int = 0
        if cantidad_int < 1:
            cantidad_int = 1
        try:
            precio = float(raw.get("precio_unitario", 0.0))
        except (TypeError, ValueError):
            precio = 0.0
        subtotal = raw.get("subtotal")
        if subtotal is None:
            subtotal = round(cantidad_int * precio, 2)
        item = {
            "codigo": _coerce_str(raw.get("codigo")),
            "descripcion": _coerce_str(raw.get("descripcion")),
            "cantidad": cantidad_int,
            "precio_unitario": round(precio, 2),
            "comentario": raw.get("comentario"),
            "subtotal": round(float(subtotal), 2),
        }
        unidad = raw.get("unidad") or raw.get("uom")
        if unidad:
            item["unidad"] = _coerce_str(unidad)
        result.append(item)
    return result


def _ensure_totals(data: dict[str, Any], fallback: float) -> float:
    try:
        stored = float(data.get("total_monto", fallback))
    except (TypeError, ValueError):
        stored = fallback
    if stored <= 0 and data.get("items"):
        subtotal = sum(item.get("subtotal", 0.0) for item in data.get("items", []))
        try:
            stored = float(subtotal)
        except (TypeError, ValueError):
            stored = fallback
    data["total_monto"] = round(stored, 2)
    return data["total_monto"]


def _assign_planner_automatically(con, centro: str, sector: str, almacen_virtual: str) -> str | None:
    """Asigna automáticamente un planificador basado en Centro, Sector y Almacén Virtual."""
    if not centro or not sector or not almacen_virtual:
        return None
    
    # Buscar asignación específica por centro, sector y almacen
    row = con.execute(
        """
        SELECT p.usuario_id
          FROM planificador_asignaciones pa
          JOIN planificadores p ON pa.planificador_id = p.usuario_id
         WHERE pa.centro = ? AND pa.sector = ? AND pa.almacen_virtual = ?
         ORDER BY pa.created_at ASC
         LIMIT 1
        """,
        (centro, sector, almacen_virtual),
    ).fetchone()
    
    if row:
        return row["usuario_id"]
    
    # Si no hay asignación específica, buscar por centro y sector
    row = con.execute(
        """
        SELECT p.usuario_id
          FROM planificador_asignaciones pa
          JOIN planificadores p ON pa.planificador_id = p.usuario_id
         WHERE pa.centro = ? AND pa.sector = ? AND pa.almacen_virtual IS NULL
         ORDER BY pa.created_at ASC
         LIMIT 1
        """,
        (centro, sector),
    ).fetchone()
    
    if row:
        return row["usuario_id"]
    
    # Si no hay asignación por centro y sector, buscar solo por centro
    row = con.execute(
        """
        SELECT p.usuario_id
          FROM planificador_asignaciones pa
          JOIN planificadores p ON pa.planificador_id = p.usuario_id
         WHERE pa.centro = ? AND pa.sector IS NULL AND pa.almacen_virtual IS NULL
         ORDER BY pa.created_at ASC
         LIMIT 1
        """,
        (centro,),
    ).fetchone()
    
    return row["usuario_id"] if row else None


def _serialize_row(row: dict[str, Any], *, detailed: bool) -> dict[str, Any]:
    data = _json_load(row.get("data_json"))
    data.setdefault("items", [])
    items = _serialize_items(data.get("items"))
    data["items"] = items
    total = _ensure_totals(data, float(row.get("total_monto") or 0.0))
    base: dict[str, Any] = {
        "id": row.get("id"),
        "status": row.get("status"),
        "centro": row.get("centro"),
        "sector": row.get("sector"),
        "justificacion": row.get("justificacion"),
        "centro_costos": row.get("centro_costos"),
        "almacen_virtual": row.get("almacen_virtual"),
        "criticidad": row.get("criticidad") or data.get("criticidad"),
        "fecha_necesidad": row.get("fecha_necesidad") or data.get("fecha_necesidad"),
        "id_usuario": row.get("id_usuario"),
        "aprobador_id": row.get("aprobador_id") or data.get("aprobador_id"),
        "planner_id": row.get("planner_id") or data.get("planner_id"),
        "total_monto": total,
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
        "notificado_at": row.get("notificado_at"),
        "data_json": data,
    }
    cancel_reason = data.get("cancel_reason")
    if cancel_reason:
        base["cancel_reason"] = cancel_reason
    if data.get("cancelled_at"):
        base["cancelled_at"] = data.get("cancelled_at")
    cancel_request = data.get("cancel_request")
    if isinstance(cancel_request, dict):
        base["cancel_request"] = cancel_request
    if detailed:
        base["items"] = items
    return base


def _load_solicitud(con, sol_id: int):
    return con.execute(
        """
        SELECT id, id_usuario, centro, sector, justificacion, centro_costos, almacen_virtual,
               data_json, status, aprobador_id, total_monto, notificado_at,
               created_at, updated_at, criticidad, fecha_necesidad, planner_id
          FROM solicitudes
         WHERE id=?
        """,
        (sol_id,),
    ).fetchone()


def _create_notification(con, destinatario: str | None, solicitud_id: int, mensaje: str) -> None:
    dest = _coerce_str(destinatario).lower()
    if not dest:
        return
    con.execute(
        "INSERT INTO notificaciones (destinatario_id, solicitud_id, mensaje, leido) VALUES (?,?,?,0)",
        (dest, solicitud_id, mensaje),
    )


def _can_view(user: dict[str, Any] | None, row: dict[str, Any]) -> bool:
    uid = _coerce_str(user.get("id_spm")) if user else ""
    if uid and uid.lower() == _coerce_str(row.get("id_usuario")).lower():
        return True
    approver = _coerce_str(row.get("aprobador_id")).lower()
    if uid and approver and uid.lower() == approver:
        return True
    planner = _coerce_str(row.get("planner_id")).lower()
    if uid and planner and uid.lower() == planner:
        return True
    if _has_role(user, "admin", "administrador", "planner", "planificador"):
        return True
    return False


def _can_decide_cancel(user: dict[str, Any] | None, row: dict[str, Any]) -> bool:
    if not user:
        return False
    uid = _coerce_str(user.get("id_spm")).lower()
    if not uid:
        return False
    approver = _coerce_str(row.get("aprobador_id")).lower()
    planner = _coerce_str(row.get("planner_id")).lower()
    if uid == approver or uid == planner:
        return True
    return _has_role(user, "admin", "administrador", "planner", "planificador")


def _can_resolve(user: dict[str, Any] | None, row: dict[str, Any]) -> bool:
    if not user:
        return False
    uid = _coerce_str(user.get("id_spm")).lower()
    if not uid:
        return False
    approver = _coerce_str(row.get("aprobador_id")).lower()
    planner = _coerce_str(row.get("planner_id")).lower()
    if uid == approver or uid == planner:
        return True
    return _has_role(user, "admin", "administrador", "aprobador", "planner", "planificador")


@bp.get("/solicitudes")
def listar_solicitudes():
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        rows = con.execute(
            """
            SELECT id, id_usuario, centro, sector, justificacion, centro_costos, almacen_virtual,
                   data_json, status, aprobador_id, total_monto, notificado_at,
                   created_at, updated_at, criticidad, fecha_necesidad, planner_id
              FROM solicitudes
             WHERE lower(id_usuario)=?
          ORDER BY datetime(created_at) DESC, id DESC
            """,
            (uid.lower(),),
        ).fetchall()
    items = [_serialize_row(row, detailed=False) for row in rows]
    return {"ok": True, "items": items, "total": len(items)}


@bp.get("/solicitudes/<int:sol_id>")
def obtener_solicitud(sol_id: int):
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        row = _load_solicitud(con, sol_id)
        if not row:
            return _json_error("NOTFOUND", "Solicitud no encontrada", 404)
        user = _fetch_user(con, uid)
        if not _can_view(user, row):
            return _json_error("FORBIDDEN", "No tienes acceso a esta solicitud", 403)
        solicitud = _serialize_row(row, detailed=True)
        # Agregar nombre del aprobador si existe
        aprobador_id = solicitud.get("aprobador_id")
        if aprobador_id:
            aprobador_user = _fetch_user(con, aprobador_id)
            if aprobador_user:
                solicitud["aprobador_nombre"] = f"{aprobador_user['nombre']} {aprobador_user['apellido']}"
        
        # Agregar nombre del planificador asignado si existe
        planner_id = solicitud.get("planner_id")
        if planner_id:
            # Buscar el nombre en la tabla planificadores
            planner_row = con.execute(
                "SELECT nombre FROM planificadores WHERE usuario_id = ?",
                (planner_id,)
            ).fetchone()
            if planner_row:
                solicitud["planner_nombre"] = planner_row["nombre"]
    return {"ok": True, "solicitud": solicitud}


def _sync_columns_from_payload(payload: dict[str, Any]) -> tuple[str, str, str, str, str, str, str]:
    return (
        payload.get("centro"),
        payload.get("sector"),
        payload.get("justificacion"),
        payload.get("centro_costos"),
        payload.get("almacen_virtual"),
        payload.get("criticidad") or "Normal",
        payload.get("fecha_necesidad"),
    )


@bp.route("/solicitudes/drafts", methods=["POST", "OPTIONS"])
def crear_borrador():
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    payload = request.get_json(force=True, silent=False) or {}
    try:
        draft_data = _parse_draft_payload(uid, payload)
    except Exception as exc:  # validation error
        return _json_error("BAD_REQUEST", str(exc), 400)
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        user = _fetch_user(con, uid)
        if not user:
            return _json_error("NOUSER", "Usuario no encontrado", 404)
        approver = _ensure_user_exists(con, _resolve_approver(con, user, 0.0))
        planner = _ensure_user_exists(con, _resolve_planner(user))
        draft_payload = {
            **draft_data,
            "items": [],
            "total_monto": 0.0,
        }
        if approver:
            draft_payload["aprobador_id"] = approver
        elif "aprobador_id" in draft_payload:
            draft_payload.pop("aprobador_id", None)
        if planner:
            draft_payload["planner_id"] = planner
        elif "planner_id" in draft_payload:
            draft_payload.pop("planner_id", None)
        data_json = json.dumps(draft_payload, ensure_ascii=False)
        centro, sector, justificacion, centro_costos, almacen_virtual, criticidad, fecha_necesidad = _sync_columns_from_payload(draft_payload)
        try:
            cur = con.execute(
                """
                INSERT INTO solicitudes (
                    id_usuario, centro, sector, justificacion, centro_costos, almacen_virtual,
                    data_json, status, aprobador_id, total_monto, criticidad, fecha_necesidad, planner_id
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    uid.lower(),
                    centro,
                    sector,
                    justificacion,
                    centro_costos,
                    almacen_virtual,
                    data_json,
                    STATUS_DRAFT,
                    approver,
                    0.0,
                    criticidad,
                    fecha_necesidad,
                    planner,
                ),
            )
            sol_id = cur.lastrowid
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo crear el borrador: {exc}", 500)
    return {"ok": True, "id": sol_id, "status": STATUS_DRAFT}


@bp.route("/solicitudes/<int:sol_id>/draft", methods=["PATCH", "OPTIONS"])
def actualizar_borrador(sol_id: int):
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    payload = request.get_json(force=True, silent=False) or {}
    try:
        draft_data = _parse_full_payload(uid, payload, expect_items=False)
    except ValueError as exc:
        return _json_error("BAD_REQUEST", str(exc), 400)
    except Exception as exc:
        return _json_error("BAD_REQUEST", str(exc), 400)
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        row = _load_solicitud(con, sol_id)
        if not row:
            return _json_error("NOTFOUND", "Solicitud no encontrada", 404)
        if _coerce_str(row.get("id_usuario")).lower() != uid.lower():
            return _json_error("FORBIDDEN", "No puedes editar este borrador", 403)
        if row.get("status") not in (STATUS_DRAFT, STATUS_CANCEL_REJECTED):
            return _json_error("INVALID_STATE", "La solicitud no está en borrador", 409)
        existing_data = _json_load(row.get("data_json"))
        existing_data.update({k: v for k, v in draft_data.items() if k != "items"})
        if draft_data.get("items"):
            existing_data["items"] = draft_data["items"]
            existing_data["total_monto"] = draft_data["total_monto"]
        
        # Recalcular aprobador si cambió el monto
        new_total = existing_data.get("total_monto", 0.0)
        old_total = row.get("total_monto", 0.0)
        if abs(new_total - old_total) > 0.01:  # Pequeña tolerancia para flotantes
            user = _fetch_user(con, uid)
            new_approver = _ensure_user_exists(con, _resolve_approver(con, user, new_total))
            if new_approver != row.get("aprobador_id"):
                existing_data["aprobador_id"] = new_approver
        
        centro, sector, justificacion, centro_costos, almacen_virtual, criticidad, fecha_necesidad = _sync_columns_from_payload(draft_data)
        data_json = json.dumps(existing_data, ensure_ascii=False)
        try:
            con.execute(
                """
                UPDATE solicitudes
                   SET centro=?, sector=?, justificacion=?, centro_costos=?, almacen_virtual=?,
                       data_json=?, total_monto=?, aprobador_id=?, criticidad=?, fecha_necesidad=?,
                       updated_at=CURRENT_TIMESTAMP
                 WHERE id=?
                """,
                (
                    centro,
                    sector,
                    justificacion,
                    centro_costos,
                    almacen_virtual,
                    data_json,
                    existing_data.get("total_monto", row.get("total_monto")),
                    existing_data.get("aprobador_id", row.get("aprobador_id")),
                    criticidad,
                    fecha_necesidad,
                    sol_id,
                ),
            )
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo guardar el borrador: {exc}", 500)
    return {"ok": True, "id": sol_id, "status": row.get("status")}


def _finalizar_solicitud(con, row: dict[str, Any], final_data: dict[str, Any], user: dict[str, Any] | None, *, is_new: bool) -> tuple[int, dict[str, Any]]:
    approver = _ensure_user_exists(
        con,
        final_data.get("aprobador_id") or _resolve_approver(con, user, final_data.get("total_monto", 0.0)),
    )
    planner = _ensure_user_exists(
        con,
        final_data.get("planner_id") or row.get("planner_id") or _resolve_planner(user),
    )
    final_payload = {**_json_load(row.get("data_json")), **final_data}
    final_payload["aprobador_id"] = approver
    if planner:
        final_payload["planner_id"] = planner
    elif "planner_id" in final_payload:
        final_payload.pop("planner_id", None)
    final_payload.pop("cancel_request", None)
    final_payload.pop("cancel_reason", None)
    final_payload.pop("cancelled_at", None)
    final_payload["total_monto"] = final_data["total_monto"]
    data_json = json.dumps(final_payload, ensure_ascii=False)
    centro, sector, justificacion, centro_costos, almacen_virtual, criticidad, fecha_necesidad = _sync_columns_from_payload(final_payload)
    now_iso = _utcnow_iso()
    if is_new:
        cur = con.execute(
            """
            INSERT INTO solicitudes (
                id_usuario, centro, sector, justificacion, centro_costos, almacen_virtual,
                data_json, status, aprobador_id, total_monto, criticidad, fecha_necesidad, planner_id, notificado_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                final_payload["id_usuario"],
                centro,
                sector,
                justificacion,
                centro_costos,
                almacen_virtual,
                data_json,
                STATUS_PENDING,
                approver,
                final_payload["total_monto"],
                criticidad,
                fecha_necesidad,
                planner,
                now_iso,
            ),
        )
        sol_id = cur.lastrowid
    else:
        con.execute(
            """
            UPDATE solicitudes
               SET centro=?, sector=?, justificacion=?, centro_costos=?, almacen_virtual=?,
                   data_json=?, status=?, aprobador_id=?, total_monto=?, criticidad=?,
                   fecha_necesidad=?, planner_id=?, notificado_at=?, updated_at=CURRENT_TIMESTAMP
             WHERE id=?
            """,
            (
                centro,
                sector,
                justificacion,
                centro_costos,
                almacen_virtual,
                data_json,
                STATUS_PENDING,
                approver,
                final_payload["total_monto"],
                criticidad,
                fecha_necesidad,
                planner,
                now_iso,
                row["id"],
            ),
        )
        sol_id = row["id"]
    if approver:
        _create_notification(con, approver, sol_id, f"Solicitud #{sol_id} pendiente de aprobación")
    return sol_id, final_payload


@bp.route("/solicitudes/<int:sol_id>", methods=["PUT", "OPTIONS"])
def finalizar_solicitud(sol_id: int):
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    payload = request.get_json(force=True, silent=False) or {}
    try:
        final_data = _parse_full_payload(uid, payload, expect_items=True)
    except ValueError as exc:
        return _json_error("BAD_REQUEST", str(exc), 400)
    except Exception as exc:
        return _json_error("BAD_REQUEST", str(exc), 400)
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        row = _load_solicitud(con, sol_id)
        if not row:
            return _json_error("NOTFOUND", "Solicitud no encontrada", 404)
        if _coerce_str(row.get("id_usuario")).lower() != uid.lower():
            return _json_error("FORBIDDEN", "No puedes finalizar esta solicitud", 403)
        if row.get("status") not in (STATUS_DRAFT, STATUS_CANCEL_REJECTED):
            return _json_error("INVALID_STATE", "La solicitud no está en borrador", 409)
        user = _fetch_user(con, uid)
        try:
            sol_id, final_payload = _finalizar_solicitud(con, row, final_data, user, is_new=False)
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo enviar la solicitud: {exc}", 500)
    return {"ok": True, "id": sol_id, "status": STATUS_PENDING, "total_monto": final_payload.get("total_monto")}


@bp.route("/solicitudes", methods=["POST", "OPTIONS"])
def crear_solicitud():
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    payload = request.get_json(force=True, silent=False) or {}
    try:
        final_data = _parse_full_payload(uid, payload, expect_items=True)
    except ValueError as exc:
        return _json_error("BAD_REQUEST", str(exc), 400)
    except Exception as exc:
        return _json_error("BAD_REQUEST", str(exc), 400)
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        dummy_row = {
            "id": None,
            "data_json": json.dumps({}, ensure_ascii=False),
            "planner_id": None,
        }
        final_data["id_usuario"] = uid.lower()
        user = _fetch_user(con, uid)
        try:
            sol_id, final_payload = _finalizar_solicitud(con, dummy_row, final_data, user, is_new=True)
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo crear la solicitud: {exc}", 500)
    return {"ok": True, "id": sol_id, "status": STATUS_PENDING, "total_monto": final_payload.get("total_monto")}


def _handle_direct_cancel(con, row: dict[str, Any], reason: str | None) -> dict[str, Any]:
    data = _json_load(row.get("data_json"))
    data["cancel_reason"] = reason or data.get("cancel_reason")
    data["cancelled_at"] = _utcnow_iso()
    data.pop("cancel_request", None)
    data_json = json.dumps(data, ensure_ascii=False)
    con.execute(
        """
        UPDATE solicitudes
           SET status=?, data_json=?, updated_at=CURRENT_TIMESTAMP
         WHERE id=?
        """,
        (STATUS_CANCELLED, data_json, row["id"]),
    )
    return data


def _create_cancel_request(row: dict[str, Any], reason: str | None, uid: str) -> dict[str, Any]:
    cancel_request = {
        "status": "pendiente",
        "reason": reason,
        "requested_at": _utcnow_iso(),
        "requested_by": uid.lower(),
    }
    return cancel_request


@bp.route("/solicitudes/<int:sol_id>/decidir", methods=["POST", "OPTIONS"])
def decidir_solicitud(sol_id: int):
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    payload = request.get_json(silent=True) or {}
    try:
        decision = BudgetIncreaseDecision(**payload)
    except Exception as exc:
        return _json_error("BAD_REQUEST", str(exc), 400)

    accion = decision.accion
    comentario = decision.comentario or None

    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        row = _load_solicitud(con, sol_id)
        if not row:
            return _json_error("NOTFOUND", "Solicitud no encontrada", 404)
        user = _fetch_user(con, uid)
        if not _can_resolve(user, row):
            return _json_error("FORBIDDEN", "No tienes permisos para esta operación", 403)
        if row.get("status") != STATUS_PENDING:
            return _json_error("INVALID_STATE", "La solicitud no está pendiente de aprobación", 409)

        decision_at = _utcnow_iso()
        data = _json_load(row.get("data_json"))
        decision_payload = {
            "status": STATUS_APPROVED if accion == "aprobar" else STATUS_REJECTED,
            "accion": accion,
            "decided_at": decision_at,
            "decided_by": uid.lower(),
        }
        if comentario:
            decision_payload["comment"] = comentario
            data["decision_comment"] = comentario
        data["decision"] = decision_payload
        data.pop("cancel_request", None)

        # Determinar status final y asignar planificador si se aprueba
        assigned_planner_id = None
        if accion == "aprobar":
            # Asignar planificador automáticamente
            centro = row.get("centro")
            sector = row.get("sector") 
            almacen_virtual = row.get("almacen_virtual")
            assigned_planner_id = _assign_planner_automatically(con, centro, sector, almacen_virtual)
            
            if assigned_planner_id:
                status_final = STATUS_IN_TREATMENT
                data["assigned_planner"] = assigned_planner_id
                message = f"Solicitud #{sol_id} aprobada y asignada al planificador"
            else:
                status_final = STATUS_APPROVED
                message = f"Solicitud #{sol_id} aprobada (sin planificador asignado)"
        else:
            status_final = STATUS_REJECTED
            message = f"Solicitud #{sol_id} rechazada"

        try:
            data_json = json.dumps(data, ensure_ascii=False)
            con.execute(
                """
                UPDATE solicitudes
                   SET status=?, data_json=?, updated_at=CURRENT_TIMESTAMP,
                       notificado_at=?, aprobador_id=?, planner_id=?
                 WHERE id=?
                """,
                (status_final, data_json, decision_at, uid.lower(), 
                 assigned_planner_id, sol_id),
            )
            owner = row.get("id_usuario")
            planner = assigned_planner_id  # Usar el planificador asignado, no el anterior
            assigned_planner = data.get("assigned_planner")
            recipients = {owner, planner, assigned_planner}
            for dest in recipients:
                if dest:
                    _create_notification(con, dest, sol_id, message)
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo registrar la decisión: {exc}", 500)

    return {"ok": True, "status": status_final, "decision": decision_payload}


@bp.route("/solicitudes/<int:sol_id>/cancel", methods=["PATCH", "OPTIONS"])
def cancelar_solicitud(sol_id: int):
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    payload = request.get_json(silent=True) or {}
    reason = _coerce_str(payload.get("reason")) or None
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        row = _load_solicitud(con, sol_id)
        if not row:
            return _json_error("NOTFOUND", "Solicitud no encontrada", 404)
        owner = _coerce_str(row.get("id_usuario")).lower()
        if owner != uid.lower():
            return _json_error("FORBIDDEN", "No puedes cancelar esta solicitud", 403)
        status = row.get("status")
        try:
            if status in (STATUS_DRAFT, STATUS_CANCEL_REJECTED):
                data = _handle_direct_cancel(con, row, reason)
                con.commit()
                return {"ok": True, "status": STATUS_CANCELLED, "cancel_reason": data.get("cancel_reason")}
            if status == STATUS_CANCELLED:
                return _json_error("INVALID_STATE", "La solicitud ya está cancelada", 409)
            data = _json_load(row.get("data_json"))
            cancel_request = _create_cancel_request(row, reason, uid)
            data["cancel_request"] = cancel_request
            if reason:
                data["cancel_reason"] = reason
            data_json = json.dumps(data, ensure_ascii=False)
            con.execute(
                """
                UPDATE solicitudes
                   SET status=?, data_json=?, updated_at=CURRENT_TIMESTAMP
                 WHERE id=?
                """,
                (STATUS_CANCEL_PENDING, data_json, sol_id),
            )
            approver = row.get("aprobador_id") or row.get("planner_id")
            _create_notification(con, approver, sol_id, f"Solicitud #{sol_id} solicita cancelación")
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo cancelar la solicitud: {exc}", 500)
    return {"ok": True, "status": STATUS_CANCEL_PENDING}


@bp.route("/solicitudes/<int:sol_id>/decidir_cancelacion", methods=["POST", "OPTIONS"])
def decidir_cancelacion(sol_id: int):
    if request.method == "OPTIONS":
        return "", 204
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)
    payload = request.get_json(force=True, silent=False) or {}
    accion = _coerce_str(payload.get("accion")).lower()
    comentario = _coerce_str(payload.get("comentario")) or None
    if accion not in {"aprobar", "rechazar"}:
        return _json_error("BAD_REQUEST", "Acción inválida", 400)
    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}
        row = _load_solicitud(con, sol_id)
        if not row:
            return _json_error("NOTFOUND", "Solicitud no encontrada", 404)
        user = _fetch_user(con, uid)
        if not _can_decide_cancel(user, row):
            return _json_error("FORBIDDEN", "No tienes permisos para esta operación", 403)
        if row.get("status") != STATUS_CANCEL_PENDING:
            return _json_error("INVALID_STATE", "La solicitud no está en cancelación pendiente", 409)
        data = _json_load(row.get("data_json"))
        cancel_request = data.get("cancel_request")
        if not isinstance(cancel_request, dict):
            cancel_request = {}
        cancel_request["decision_at"] = _utcnow_iso()
        cancel_request["decision_by"] = uid.lower()
        if comentario:
            cancel_request["decision_comment"] = comentario
        owner = row.get("id_usuario")
        try:
            if accion == "aprobar":
                cancel_request["status"] = "aprobada"
                data["cancel_request"] = cancel_request
                data["cancelled_at"] = cancel_request["decision_at"]
                data_json = json.dumps(data, ensure_ascii=False)
                con.execute(
                    """
                    UPDATE solicitudes
                       SET status=?, data_json=?, updated_at=CURRENT_TIMESTAMP
                     WHERE id=?
                    """,
                    (STATUS_CANCELLED, data_json, sol_id),
                )
                _create_notification(con, owner, sol_id, f"Solicitud #{sol_id} cancelada")
                result_status = STATUS_CANCELLED
            else:
                cancel_request["status"] = "rechazada"
                data["cancel_request"] = cancel_request
                data_json = json.dumps(data, ensure_ascii=False)
                con.execute(
                    """
                    UPDATE solicitudes
                       SET status=?, data_json=?, updated_at=CURRENT_TIMESTAMP
                     WHERE id=?
                    """,
                    (STATUS_CANCEL_REJECTED, data_json, sol_id),
                )
                _create_notification(con, owner, sol_id, f"Solicitud #{sol_id}: cancelación rechazada")
                result_status = STATUS_CANCEL_REJECTED
            con.commit()
        except Exception as exc:
            con.rollback()
            return _json_error("DB_ERROR", f"No se pudo registrar la decisión: {exc}", 500)
    return {"ok": True, "status": result_status, "accion": accion}


@bp.get("/solicitudes/export/excel")
def export_solicitudes_excel():
    """Exportar todas las solicitudes del usuario autenticado a Excel"""
    user_id = _require_auth()
    if not user_id:
        return _json_error("UNAUTHORIZED", "Autenticación requerida", 401)

    try:
        with get_connection() as con:
            # Obtener todas las solicitudes del usuario
            solicitudes = con.execute("""
                SELECT id, centro, sector, centro_costos, almacen_virtual, criticidad,
                       fecha_necesidad, justificacion, status, created_at, updated_at,
                       total_monto, aprobador_id, data_json
                FROM solicitudes
                WHERE id_usuario = ?
                ORDER BY created_at DESC
            """, (user_id,)).fetchall()

            if not solicitudes:
                return _json_error("NO_DATA", "No hay solicitudes para exportar", 404)

            # Crear workbook de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Mis Solicitudes"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center")

            # Headers principales
            headers = [
                "ID", "Centro", "Sector", "Centro de Costos", "Almacén Virtual",
                "Criticidad", "Fecha Necesidad", "Justificación", "Estado",
                "Fecha Creación", "Última Actualización", "Total Estimado", "Aprobador"
            ]

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment

            # Datos de solicitudes
            current_row = 2
            for sol in solicitudes:
                try:
                    # Asegurarse de que sol sea una tupla/lista
                    if not isinstance(sol, (tuple, list)):
                        continue
                    
                    # Datos principales de la solicitud
                    ws.cell(row=current_row, column=1, value=sol[0])  # ID
                    ws.cell(row=current_row, column=2, value=sol[1] or "")  # Centro
                    ws.cell(row=current_row, column=3, value=sol[2] or "")  # Sector
                    ws.cell(row=current_row, column=4, value=sol[3] or "")  # Centro Costos
                    ws.cell(row=current_row, column=5, value=sol[4] or "")  # Almacén Virtual
                    ws.cell(row=current_row, column=6, value=sol[5] or "")  # Criticidad
                    ws.cell(row=current_row, column=7, value=sol[6] or "")  # Fecha Necesidad
                    ws.cell(row=current_row, column=8, value=sol[7] or "")  # Justificación
                    ws.cell(row=current_row, column=9, value=sol[8] or "")  # Estado
                    ws.cell(row=current_row, column=10, value=sol[9] or "")  # Fecha Creación
                    ws.cell(row=current_row, column=11, value=sol[10] or "")  # Última Actualización
                    ws.cell(row=current_row, column=12, value=sol[11] or 0)  # Total Estimado

                    # Obtener nombre del aprobador si existe
                    aprobador_name = ""
                    if len(sol) > 12 and sol[12]:  # aprobador_id
                        try:
                            aprobador = con.execute("SELECT nombre FROM usuarios WHERE id = ?", (sol[12],)).fetchone()
                            if aprobador:
                                aprobador_name = aprobador[0] or ""
                        except Exception:
                            aprobador_name = ""
                    ws.cell(row=current_row, column=13, value=aprobador_name)

                    current_row += 1

                    # Agregar items de la solicitud en filas separadas
                    data_json = sol[13] if len(sol) > 13 else None
                    if data_json and isinstance(data_json, str):
                        try:
                            data = json.loads(data_json)
                            items = data.get('items', [])
                            if items:
                                # Agregar fila de separación
                                ws.cell(row=current_row, column=1, value=f"Items de Solicitud #{sol[0]}")
                                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=13)
                                current_row += 1

                                # Headers de items
                                item_headers = ["Código", "Descripción", "Unidad", "Precio Unitario", "Cantidad", "Subtotal"]
                                for col_num, header in enumerate(item_headers, 1):
                                    cell = ws.cell(row=current_row, column=col_num, value=header)
                                    cell.font = Font(bold=True)
                                current_row += 1

                                # Datos de items
                                for item in items:
                                    if isinstance(item, dict):
                                        ws.cell(row=current_row, column=1, value=item.get('codigo', ''))
                                        ws.cell(row=current_row, column=2, value=item.get('descripcion', ''))
                                        ws.cell(row=current_row, column=3, value=item.get('unidad', ''))
                                        ws.cell(row=current_row, column=4, value=item.get('precio_unitario', 0))
                                        ws.cell(row=current_row, column=5, value=item.get('cantidad', 0))
                                        ws.cell(row=current_row, column=6, value=item.get('subtotal', 0))
                                        current_row += 1
                        except (json.JSONDecodeError, KeyError, TypeError):
                            # Si hay problemas con el JSON, continuar sin items
                            pass
                except Exception as row_error:
                    # Si hay error en una fila específica, continuar con la siguiente
                    print(f"Error procesando solicitud: {row_error}")
                    continue

            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres de ancho
                ws.column_dimensions[column_letter].width = adjusted_width

            # Guardar en buffer
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return send_file(
                buffer,
                as_attachment=True,
                download_name=f"mis_solicitudes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as exc:
        print(f"Error en export_solicitudes_excel: {exc}")
        import traceback
        traceback.print_exc()
        return _json_error("EXPORT_ERROR", f"Error al exportar a Excel: {exc}", 500)


@bp.get("/solicitudes/export/pdf")
def export_solicitudes_pdf():
    """Exportar todas las solicitudes del usuario autenticado a PDF"""
    user_id = _require_auth()
    if not user_id:
        return _json_error("UNAUTHORIZED", "Autenticación requerida", 401)

    try:
        with get_connection() as con:
            # Obtener todas las solicitudes del usuario
            solicitudes = con.execute("""
                SELECT id, centro, sector, centro_costos, almacen_virtual, criticidad,
                       fecha_necesidad, justificacion, status, created_at, updated_at,
                       total_monto, aprobador_id, data_json
                FROM solicitudes
                WHERE id_usuario = ?
                ORDER BY created_at DESC
            """, (user_id,)).fetchall()

            if not solicitudes:
                return _json_error("NO_DATA", "No hay solicitudes para exportar", 404)

            # Crear buffer para el PDF
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()

            # Estilos personalizados
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Centrado
            )

            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=20,
                alignment=0  # Izquierda
            )

            normal_style = styles['Normal']

            story = []

            # Título del documento
            story.append(Paragraph("Mis Solicitudes - SPM", title_style))
            story.append(Spacer(1, 12))
            story.append(Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", normal_style))
            story.append(Spacer(1, 20))

            for sol in solicitudes:
                # Información de la solicitud
                story.append(Paragraph(f"Solicitud #{sol[0]}", subtitle_style))

                solicitud_data = [
                    ["Centro:", sol[1] or "-"],
                    ["Sector:", sol[2] or "-"],
                    ["Centro de Costos:", sol[3] or "-"],
                    ["Almacén Virtual:", sol[4] or "-"],
                    ["Criticidad:", sol[5] or "-"],
                    ["Fecha Necesidad:", sol[6] or "-"],
                    ["Estado:", sol[8] or "-"],
                    ["Fecha Creación:", sol[9] or "-"],
                    ["Total Estimado:", f"${sol[11] or 0:.2f}" if sol[11] else "-"],
                ]

                # Obtener nombre del aprobador
                if sol[12]:
                    aprobador = con.execute("SELECT nombre FROM usuarios WHERE id = ?", (sol[12],)).fetchone()
                    if aprobador:
                        solicitud_data.append(["Aprobador:", aprobador[0]])

                solicitud_table = Table(solicitud_data, colWidths=[100, 300])
                solicitud_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('BACKGROUND', (1, 0), (1, -1), colors.white),
                ]))
                story.append(solicitud_table)
                story.append(Spacer(1, 12))

                # Justificación
                if sol[7]:
                    story.append(Paragraph("Justificación:", styles['Heading3']))
                    story.append(Paragraph(sol[7], normal_style))
                    story.append(Spacer(1, 12))

                # Items de la solicitud
                data_json = sol[13]
                if data_json and isinstance(data_json, str):
                    try:
                        data = json.loads(data_json)
                        items = data.get('items', [])
                        if items:
                            story.append(Paragraph("Items Solicitados:", styles['Heading3']))

                            item_data = [["Código", "Descripción", "Unidad", "Precio Unit.", "Cantidad", "Subtotal"]]
                            for item in items:
                                item_data.append([
                                    item.get('codigo', ''),
                                    item.get('descripcion', ''),
                                    item.get('unidad', ''),
                                    f"${item.get('precio_unitario', 0):.2f}",
                                    str(item.get('cantidad', 0)),
                                    f"${item.get('subtotal', 0):.2f}"
                                ])

                            item_table = Table(item_data, colWidths=[60, 150, 50, 70, 60, 70])
                            item_table.setStyle(TableStyle([
                                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                ('FONTSIZE', (0, 0), (-1, 0), 10),
                                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                                ('ALIGN', (3, 1), (5, -1), 'RIGHT'),
                                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                                ('FONTSIZE', (0, 1), (-1, -1), 9),
                                ('GRID', (0, 0), (-1, -1), 1, colors.black)
                            ]))
                            story.append(item_table)
                            story.append(Spacer(1, 20))
                    except json.JSONDecodeError:
                        pass

                # Separador entre solicitudes
                story.append(Paragraph("-" * 80, normal_style))
                story.append(Spacer(1, 20))

            # Generar PDF
            doc.build(story)
            buffer.seek(0)

            return send_file(
                buffer,
                as_attachment=True,
                download_name=f"mis_solicitudes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                mimetype="application/pdf"
            )

    except Exception as exc:
        return _json_error("EXPORT_ERROR", f"Error al exportar a PDF: {exc}", 500)


@bp.get("/solicitudes/equipo")
def listar_solicitudes_equipo():
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)

    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}

        # Verificar que el usuario tenga permisos para ver solicitudes del equipo
        user = _fetch_user(con, uid)
        if not user:
            return _json_error("NOTFOUND", "Usuario no encontrado", 404)

        # Verificar si el usuario es jefe o gerente (tiene personas reportando a él)
        team_user_ids = []
        user_id = user.get("id_spm") or user.get("id")

        # Buscar usuarios que reporten a este usuario como jefe, gerente1 o gerente2
        team_rows = con.execute(
            """
            SELECT id_spm, id
              FROM usuarios
             WHERE lower(jefe) = ? OR lower(gerente1) = ? OR lower(gerente2) = ?
            """,
            (user_id.lower(), user_id.lower(), user_id.lower())
        ).fetchall()

        if not team_rows:
            # Si no tiene equipo, devolver lista vacía
            return {"ok": True, "items": [], "total": 0}

        # Recopilar IDs de los miembros del equipo
        for team_row in team_rows:
            team_uid = team_row.get("id_spm") or team_row.get("id")
            if team_uid:
                team_user_ids.append(team_uid.lower())

        if not team_user_ids:
            return {"ok": True, "items": [], "total": 0}

        # Crear placeholders para la consulta IN
        placeholders = ','.join('?' for _ in team_user_ids)

        # Obtener solicitudes de los miembros del equipo
        rows = con.execute(
            f"""
            SELECT s.id, s.id_usuario, s.centro, s.sector, s.justificacion, s.centro_costos, s.almacen_virtual,
                   s.data_json, s.status, s.aprobador_id, s.total_monto, s.notificado_at,
                   s.created_at, s.updated_at, s.criticidad, s.fecha_necesidad, s.planner_id,
                   u.nombre as solicitante_nombre, u.apellido as solicitante_apellido
              FROM solicitudes s
              JOIN usuarios u ON lower(s.id_usuario) = lower(u.id_spm) OR lower(s.id_usuario) = lower(u.id)
             WHERE lower(s.id_usuario) IN ({placeholders})
          ORDER BY datetime(s.created_at) DESC, s.id DESC
            """,
            team_user_ids
        ).fetchall()

    items = []
    for row in rows:
        item = _serialize_row(row, detailed=False)
        # Agregar nombre del solicitante
        if row.get("solicitante_nombre") and row.get("solicitante_apellido"):
            item["solicitante"] = f"{row['solicitante_nombre']} {row['solicitante_apellido']}"
        items.append(item)

    return {"ok": True, "items": items, "total": len(items)}


@bp.get("/reportes/estadisticas")
def obtener_estadisticas():
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)

    with get_connection() as con:
        con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}

        # Estadísticas generales
        stats = con.execute(
            """
            SELECT
                COUNT(*) as total_solicitudes,
                COUNT(CASE WHEN status IN ('approved', 'en_tratamiento', 'completed') THEN 1 END) as solicitudes_activas,
                COUNT(CASE WHEN status = 'completed' THEN 1 END) as solicitudes_completadas,
                SUM(CASE WHEN status IN ('approved', 'en_tratamiento', 'completed') THEN total_monto ELSE 0 END) as monto_total
            FROM solicitudes
            """
        ).fetchone()

        # Solicitudes por estado
        estado_stats = con.execute(
            """
            SELECT status, COUNT(*) as cantidad
            FROM solicitudes
            GROUP BY status
            ORDER BY cantidad DESC
            """
        ).fetchall()

        # Solicitudes por centro
        centro_stats = con.execute(
            """
            SELECT centro, COUNT(*) as cantidad
            FROM solicitudes
            GROUP BY centro
            ORDER BY cantidad DESC
            LIMIT 10
            """
        ).fetchall()

        # Tendencia mensual (últimos 12 meses)
        mensual_stats = con.execute(
            """
            SELECT
                strftime('%Y-%m', created_at) as mes,
                COUNT(*) as cantidad
            FROM solicitudes
            WHERE created_at >= date('now', '-12 months')
            GROUP BY strftime('%Y-%m', created_at)
            ORDER BY mes
            """
        ).fetchall()

    return {
        "ok": True,
        "estadisticas": {
            "total_solicitudes": stats["total_solicitudes"] or 0,
            "solicitudes_activas": stats["solicitudes_activas"] or 0,
            "solicitudes_completadas": stats["solicitudes_completadas"] or 0,
            "monto_total": float(stats["monto_total"] or 0)
        },
        "por_estado": [{"estado": row["status"], "cantidad": row["cantidad"]} for row in estado_stats],
        "por_centro": [{"centro": row["centro"], "cantidad": row["cantidad"]} for row in centro_stats],
        "tendencia_mensual": [{"mes": row["mes"], "cantidad": row["cantidad"]} for row in mensual_stats]
    }


@bp.get("/reportes/exportar")
def exportar_reporte():
    uid = _require_auth()
    if not uid:
        return _json_error("NOAUTH", "No autenticado", 401)

    tipo = request.args.get("tipo", "solicitudes")
    formato = request.args.get("formato", "excel")
    fecha_desde = request.args.get("fecha_desde")
    fecha_hasta = request.args.get("fecha_hasta")

    try:
        with get_connection() as con:
            con.row_factory = lambda cursor, row: {col[0]: row[idx] for idx, col in enumerate(cursor.description)}

            # Construir consulta base
            query = """
                SELECT s.id, s.id_usuario, s.centro, s.sector, s.justificacion, s.centro_costos,
                       s.almacen_virtual, s.status, s.aprobador_id, s.total_monto, s.created_at,
                       s.updated_at, s.criticidad, s.fecha_necesidad, s.planner_id,
                       u.nombre, u.apellido, p.nombre as planner_nombre
                FROM solicitudes s
                JOIN usuarios u ON lower(s.id_usuario) = lower(u.id_spm) OR lower(s.id_usuario) = lower(u.id)
                LEFT JOIN planificadores p ON s.planner_id = p.usuario_id
            """
            params = []

            # Agregar filtros de fecha
            where_clauses = []
            if fecha_desde:
                where_clauses.append("s.created_at >= ?")
                params.append(fecha_desde)
            if fecha_hasta:
                where_clauses.append("s.created_at <= ?")
                params.append(fecha_hasta + " 23:59:59")

            if where_clauses:
                query += " WHERE " + " AND ".join(where_clauses)

            query += " ORDER BY s.created_at DESC"

            rows = con.execute(query, params).fetchall()

        if formato == "excel":
            return _exportar_excel(rows)
        elif formato == "pdf":
            return _exportar_pdf(rows)
        else:
            return _json_error("INVALID_FORMAT", "Formato no soportado", 400)

    except Exception as exc:
        return _json_error("EXPORT_ERROR", f"Error al generar reporte: {exc}", 500)


def _exportar_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes"

    # Headers
    headers = [
        "ID", "Solicitante", "Centro", "Sector", "Estado", "Monto Total",
        "Fecha Creación", "Fecha Actualización", "Planificador", "Críticidad"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row["id"])
        ws.cell(row=row_idx, column=2, value=f"{row['nombre']} {row['apellido']}")
        ws.cell(row=row_idx, column=3, value=row["centro"])
        ws.cell(row=row_idx, column=4, value=row["sector"])
        ws.cell(row=row_idx, column=5, value=row["status"])
        ws.cell(row=row_idx, column=6, value=row["total_monto"])
        ws.cell(row=row_idx, column=7, value=row["created_at"])
        ws.cell(row=row_idx, column=8, value=row["updated_at"])
        ws.cell(row=row_idx, column=9, value=row.get("planner_nombre", ""))
        ws.cell(row=row_idx, column=10, value=row["criticidad"])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"reporte_solicitudes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def _exportar_pdf(rows):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
    )
    normal_style = styles['Normal']

    # Título
    story.append(Paragraph("Reporte de Solicitudes", title_style))
    story.append(Spacer(1, 20))

    # Tabla de datos
    data = [["ID", "Solicitante", "Centro", "Estado", "Monto", "Fecha"]]

    for row in rows:
        data.append([
            str(row["id"]),
            f"{row['nombre']} {row['apellido']}",
            row["centro"] or "",
            row["status"] or "",
            f"${row['total_monto']:.2f}" if row["total_monto"] else "$0.00",
            row["created_at"][:10] if row["created_at"] else ""
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(table)

    doc.build(story)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"reporte_solicitudes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
        mimetype="application/pdf"
    )

